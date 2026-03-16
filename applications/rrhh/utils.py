import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def _estilo_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="1e4d78")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def exportar_empleados_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    headers = ['Código', 'Apellidos', 'Nombres', 'DNI', 'Cargo', 'Departamento',
               'Fecha Ingreso', 'Sueldo Base', 'Estado']
    ws.append(headers)
    _estilo_header(ws, 1, len(headers))

    for e in queryset:
        ws.append([
            e.codigo, e.apellidos, e.nombres, e.dni,
            e.cargo.nombre, e.departamento.nombre,
            e.fecha_ingreso.strftime('%d/%m/%Y'),
            float(e.sueldo_base),
            e.get_estado_display(),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="empleados.xlsx"'
    wb.save(response)
    return response


def exportar_asistencia_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    headers = ['Empleado', 'Fecha', 'Hora Entrada', 'Hora Salida', 'Estado', 'Observación']
    ws.append(headers)
    _estilo_header(ws, 1, len(headers))

    for a in queryset:
        ws.append([
            str(a.empleado),
            a.fecha.strftime('%d/%m/%Y'),
            str(a.hora_entrada) if a.hora_entrada else '—',
            str(a.hora_salida) if a.hora_salida else '—',
            a.get_estado_display(),
            a.observacion,
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="asistencia.xlsx"'
    wb.save(response)
    return response


def exportar_empleados_pdf(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=16, textColor=colors.HexColor('#1e4d78'))
    elements.append(Paragraph("Reporte de Empleados", title_style))
    elements.append(Spacer(1, 0.5*cm))

    data = [['Código', 'Apellidos y Nombres', 'DNI', 'Cargo', 'Departamento', 'F. Ingreso', 'Estado']]
    for e in queryset:
        data.append([
            e.codigo,
            f"{e.apellidos}, {e.nombres}"[:35],
            e.dni,
            e.cargo.nombre[:20],
            e.departamento.nombre[:20],
            e.fecha_ingreso.strftime('%d/%m/%Y'),
            e.get_estado_display(),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e4d78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef2f7')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="empleados.pdf"'
    return response
