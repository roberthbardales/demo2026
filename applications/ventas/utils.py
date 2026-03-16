import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def _estilo_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="1a3c5e")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def exportar_pedidos_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    headers = ['Numero', 'Cliente', 'Vendedor', 'Fecha', 'Estado', 'Subtotal', 'IGV', 'Total']
    ws.append(headers)
    _estilo_header(ws, 1, len(headers))

    estados = {'P': 'Pendiente', 'C': 'Confirmado', 'E': 'Entregado', 'A': 'Anulado'}
    for p in queryset:
        ws.append([
            p.numero,
            str(p.cliente),
            p.vendedor.full_name,
            p.fecha.strftime('%d/%m/%Y'),
            estados.get(p.estado, p.estado),
            float(p.subtotal),
            float(p.igv),
            float(p.total),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
    wb.save(response)
    return response


def exportar_comprobantes_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    headers = ['Comprobante', 'Cliente', 'Vendedor', 'Fecha', 'Subtotal', 'IGV', 'Total', 'Estado']
    ws.append(headers)
    _estilo_header(ws, 1, len(headers))

    for c in queryset:
        ws.append([
            str(c),
            str(c.cliente),
            c.vendedor.full_name,
            c.fecha_emision.strftime('%d/%m/%Y'),
            float(c.subtotal),
            float(c.igv),
            float(c.total),
            'Emitido' if c.estado == 'E' else 'Anulado',
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="comprobantes.xlsx"'
    wb.save(response)
    return response


def exportar_comprobante_pdf(comprobante):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Encabezado
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=18, textColor=colors.HexColor('#1a3c5e'),
                                 spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                               fontSize=10, textColor=colors.grey)
    normal = styles['Normal']

    tipo = comprobante.get_tipo_display().upper()
    elements.append(Paragraph(f"{tipo}", title_style))
    elements.append(Paragraph(f"{comprobante.numero_completo}", sub_style))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a3c5e')))
    elements.append(Spacer(1, 0.5*cm))

    # Datos cliente y fecha
    info_data = [
        ['Cliente:', str(comprobante.cliente), 'Fecha:', comprobante.fecha_emision.strftime('%d/%m/%Y')],
        ['Vendedor:', comprobante.vendedor.full_name, 'Estado:', comprobante.get_estado_display()],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 8*cm, 3*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*cm))

    # Detalle de productos
    data = [['Descripcion', 'Cant.', 'P. Unit.', 'Desc.%', 'Subtotal']]
    for det in comprobante.detalles.all():
        data.append([
            det.descripcion[:40],
            str(det.cantidad),
            f"S/ {det.precio_unitario:.2f}",
            f"{det.descuento}%",
            f"S/ {det.subtotal:.2f}",
        ])

    table = Table(data, colWidths=[9*cm, 2*cm, 3*cm, 2*cm, 3*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c5e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))

    # Totales
    totales_data = [
        ['', '', 'Subtotal:', f"S/ {comprobante.subtotal:.2f}"],
        ['', '', 'IGV (18%):', f"S/ {comprobante.igv:.2f}"],
        ['', '', 'TOTAL:', f"S/ {comprobante.total:.2f}"],
    ]
    totales_table = Table(totales_data, colWidths=[5*cm, 4*cm, 4*cm, 4*cm])
    totales_table.setStyle(TableStyle([
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 2), (3, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (2, 2), (3, 2), 12),
        ('TEXTCOLOR', (2, 2), (3, 2), colors.HexColor('#1a3c5e')),
    ]))
    elements.append(totales_table)

    if comprobante.observaciones:
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"Observaciones: {comprobante.observaciones}", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{comprobante.numero_completo}.pdf"'
    return response