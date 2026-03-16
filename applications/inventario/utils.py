import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ─── EXCEL ────────────────────────────────────────────────

def _estilo_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="1a3c5e")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def exportar_productos_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ['Código', 'Nombre', 'Categoría', 'Unidad', 'Proveedor',
               'P. Compra', 'P. Venta', 'Stock Actual', 'Stock Mín.', 'Estado']
    ws.append(headers)
    _estilo_header(ws, 1, len(headers))

    for p in queryset:
        ws.append([
            p.codigo, p.nombre,
            p.categoria.nombre,
            p.unidad_medida.abreviatura,
            p.proveedor.nombre if p.proveedor else '—',
            float(p.precio_compra), float(p.precio_venta),
            float(p.stock_actual), float(p.stock_minimo),
            'Activo' if p.activo else 'Inactivo',
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response


def exportar_movimientos_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = ['Fecha', 'Tipo', 'Producto', 'Almacén', 'Cantidad', 'P. Unitario', 'Total', 'Motivo', 'Usuario']
    ws.append(headers)
    _estilo_header(ws, 1, len(headers))

    for m in queryset:
        ws.append([
            m.fecha.strftime('%d/%m/%Y %H:%M'),
            m.get_tipo_display(),
            str(m.producto),
            str(m.almacen),
            float(m.cantidad),
            float(m.precio_unitario),
            float(m.total),
            m.motivo,
            str(m.usuario),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="movimientos.xlsx"'
    wb.save(response)
    return response


# ─── PDF ─────────────────────────────────────────────────

def exportar_productos_pdf(queryset):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=16, textColor=colors.HexColor('#1a3c5e'))
    elements.append(Paragraph("Reporte de Productos", title_style))
    elements.append(Spacer(1, 0.5*cm))

    data = [['Código', 'Nombre', 'Categoría', 'Unidad', 'P.Compra', 'P.Venta', 'Stock', 'Estado']]
    for p in queryset:
        data.append([
            p.codigo, p.nombre[:30],
            p.categoria.nombre[:20],
            p.unidad_medida.abreviatura,
            f"S/ {p.precio_compra:.2f}",
            f"S/ {p.precio_venta:.2f}",
            str(p.stock_actual),
            'Activo' if p.activo else 'Inactivo',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c5e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    return response
