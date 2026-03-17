from django.urls import reverse_lazy, reverse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, HttpResponse
from django.http import HttpResponseRedirect
from django.views.generic import View, ListView,TemplateView
from django.views.generic.edit import FormView

from .forms import UserRegisterForm, LoginForm, UpdatePasswordForm,UserUpdateForm
from .models import User
from .mixins import SuperusuarioMixin, UsuarioListaMixin
from django.views.generic import View, ListView, TemplateView, UpdateView
from django.contrib import messages

#exportar pdf y excel
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


class UserRegisterView(SuperusuarioMixin, FormView):
    """Solo superusuario puede registrar nuevos usuarios."""
    template_name = 'users/register.html'
    form_class = UserRegisterForm
    success_url = reverse_lazy('users_app:user-lista')

    def form_valid(self, form):
        User.objects.create_user(
            form.cleaned_data['email'],
            form.cleaned_data['password1'],
            full_name=form.cleaned_data['full_name'],
            occupation=form.cleaned_data['occupation'],
            genero=form.cleaned_data['genero'],
            date_birth=form.cleaned_data['date_birth'],
            avatar=form.cleaned_data.get('avatar'),
        )
        return super().form_valid(form)


class LoginUser(FormView):
    template_name = 'users/login.html'
    form_class = LoginForm
    success_url = reverse_lazy('home_app:dashboard')  # fallback por defecto

    def form_valid(self, form):
        user = authenticate(
            email=form.cleaned_data['email'],
            password=form.cleaned_data['password']
        )
        login(self.request, user)

        # Redirigir según rol
        if user.is_superuser or user.occupation == User.ADMIN:
            return HttpResponseRedirect(reverse('users_app:user-lista'))
        elif user.occupation == User.RRHH:
            return HttpResponseRedirect(reverse('rrhh_app:dashboard'))
        elif user.occupation == User.INVENTARIO:
            return HttpResponseRedirect(reverse('inventario_app:dashboard'))
        elif user.occupation == User.VENTAS:          # ← AGREGA ESTA LÍNEA
            return HttpResponseRedirect(reverse('ventas_app:dashboard'))   # ← Y ESTA
        elif user.occupation == User.EMPLEADO:
            return HttpResponseRedirect(reverse('home_app:inicio'))
        else:
            return HttpResponseRedirect(reverse('home_app:inicio'))


class LogoutView(View):
    def get(self, request, *args, **kwargs):
        logout(request)
        return HttpResponseRedirect(reverse('users_app:user-login'))


class UpdatePasswordView(LoginRequiredMixin, FormView):
    """Cualquier usuario logueado puede cambiar su contraseña."""
    template_name = 'users/cambiar_contraseña.html'
    form_class = UpdatePasswordForm
    success_url = reverse_lazy('users_app:user-login')
    login_url = reverse_lazy('users_app:user-login')

    def form_valid(self, form):
        usuario = self.request.user
        user = authenticate(
            email=usuario.email,
            password=form.cleaned_data['password1']
        )
        if user:
            usuario.set_password(form.cleaned_data['password2'])
            usuario.save()
        logout(self.request)
        return super().form_valid(form)


class UserListView(UsuarioListaMixin, ListView):
    """Admin y RRHH pueden ver la lista de usuarios."""
    template_name = 'users/lista_usuarios.html'
    context_object_name = 'usuarios'

    def get_queryset(self):
        return User.objects.usuarios_sistema()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        ctx['total_admin']      = qs.filter(occupation=User.ADMIN).count()
        ctx['total_rrhh']       = qs.filter(occupation=User.RRHH).count()
        ctx['total_inventario'] = qs.filter(occupation=User.INVENTARIO).count()
        ctx['total_empleado']   = qs.filter(occupation=User.EMPLEADO).count()
        return ctx

class UserUpdateView(SuperusuarioMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = 'users/user_update.html'
    success_url = reverse_lazy('users_app:user-lista')

    def form_valid(self, form):
        messages.success(self.request, "Usuario actualizado correctamente.")
        return super().form_valid(form)


# ─── REPORTES ────────────────────────────────────────────

class UserReporteExcelView(UsuarioListaMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        qs = User.objects.usuarios_sistema()
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        headers = ['Nombre Completo', 'Email', 'Rol', 'Genero', 'Fecha Nac.', 'Estado']
        ws.append(headers)

        fill = PatternFill("solid", fgColor="1a3c5e")
        font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

        roles = {'0': 'Administrador', '1': 'RRHH', '2': 'Inventario', '3': 'Empleado'}
        generos = {'M': 'Masculino', 'F': 'Femenino', 'O': 'Otro'}

        for u in qs:
            ws.append([
                u.full_name,
                u.email,
                roles.get(u.occupation, u.occupation),
                generos.get(u.genero, '-'),
                u.date_birth.strftime('%d/%m/%Y') if u.date_birth else '-',
                'Activo' if u.is_active else 'Inactivo',
            ])

        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
        wb.save(response)
        return response


class UserReportePDFView(UsuarioListaMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        qs = User.objects.usuarios_sistema()
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=1*cm, leftMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                     fontSize=16, textColor=colors.HexColor('#1a3c5e'))
        elements.append(Paragraph("Reporte de Usuarios del Sistema", title_style))
        elements.append(Spacer(1, 0.5*cm))

        roles = {'0': 'Administrador', '1': 'RRHH', '2': 'Inventario', '3': 'Empleado'}
        data = [['Nombre Completo', 'Email', 'Rol', 'Estado']]
        for u in qs:
            data.append([
                u.full_name,
                u.email,
                roles.get(u.occupation, u.occupation),
                'Activo' if u.is_active else 'Inactivo',
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c5e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="usuarios.pdf"'
        return response
